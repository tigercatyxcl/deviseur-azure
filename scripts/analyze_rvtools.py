#!/usr/bin/env python3
"""
Batch sizing from a VMware RVTools export.

Reads an RVTools ``.xlsx`` (the ``vInfo`` sheet), and for every VM maps its
on-prem allocation (vCPU, RAM, provisioned disk) to the cheapest Azure flavor
that meets-or-exceeds it (lift-and-shift / right-size-not). Pulls live prices
and prints a per-VM table plus rollup totals for a target region.

Usage:
    python3 scripts/analyze_rvtools.py inventory.xlsx --region francecentral
    python3 scripts/analyze_rvtools.py inventory.xlsx --include-poweredoff \
        --disk-type standard-ssd --output

Notes:
  * RVTools is a point-in-time *allocation* snapshot, not performance data, and
    contains NO Azure region — region is a target you choose (default
    francecentral). For performance-based right-sizing use Azure Migrate.
  * Provisioned storage is modelled as a single managed disk per VM, billed at
    the next tier up. Real migrations split OS/data disks.
"""

import argparse
import io
import os
import re
import sys
from contextlib import redirect_stdout
from datetime import date
from typing import Any, Dict, List, Optional, Tuple

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from azure_lib import (  # noqa: E402
    load_catalog, load_disk_tiers, pick_flavor,
    query_vm_prices, organize_vm_prices,
    disk_tier_for_size, query_disk_price,
    currency_symbol, HOURS_PER_MONTH, HOURS_PER_YEAR,
)

FAMILY_LABEL = {"B": "Burstable", "D": "General", "E": "Memory", "F": "Compute"}


# ── RVTools parsing ────────────────────────────────────────────────────────

def _to_number(value: Any) -> Optional[float]:
    """Parse a cell into a float, tolerating commas/spaces/blanks."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip().replace(",", "").replace(" ", "")
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _unit_to_gib(value: Optional[float], header: str) -> Optional[float]:
    """Convert a capacity cell to GiB based on the unit hinted in its header.

    RVTools reports memory and disk in MiB by default; headers may also say
    GiB/GB or TiB/TB. Anything unrecognised is assumed MiB.
    """
    if value is None:
        return None
    h = header.lower()
    if "tib" in h or re.search(r"\btb\b", h):
        return value * 1024
    if "gib" in h or re.search(r"\bgb\b", h):
        return value
    return value / 1024  # MiB (default)


def _find_column(headers: List[str], *, exact: List[str], contains: List[str]) -> Optional[int]:
    """Locate a column by header. Prefer an exact (lowercased) match, then the
    first header containing any of the substrings."""
    low = [h.strip().lower() for h in headers]
    for want in exact:
        if want in low:
            return low.index(want)
    for i, h in enumerate(low):
        if any(sub in h for sub in contains):
            return i
    return None


def _pick_sheet(wb, requested: Optional[str]):
    """Return the worksheet to parse: the requested one, else 'vInfo', else the
    first sheet that has both a CPU and a Memory column."""
    if requested:
        for ws in wb.worksheets:
            if ws.title.strip().lower() == requested.strip().lower():
                return ws
        raise ValueError(f"Sheet '{requested}' not found. Sheets: {wb.sheetnames}")
    for ws in wb.worksheets:
        if ws.title.strip().lower() == "vinfo":
            return ws
    for ws in wb.worksheets:
        headers = [str(c.value or "") for c in next(ws.iter_rows(max_row=1))]
        if (_find_column(headers, exact=["cpus"], contains=["cpu"]) is not None
                and _find_column(headers, exact=["memory"], contains=["memory"]) is not None):
            return ws
    raise ValueError(f"No vInfo-like sheet found. Sheets: {wb.sheetnames}")


def parse_rvtools(path: str, sheet: Optional[str],
                  include_off: bool, include_templates: bool) -> Tuple[List[Dict[str, Any]], Dict[str, int]]:
    """Read VM rows from an RVTools xlsx. Returns (vms, skipped_counts)."""
    import openpyxl
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = _pick_sheet(wb, sheet)

    rows = ws.iter_rows(values_only=True)
    headers = [str(h or "") for h in next(rows)]

    col_name = _find_column(headers, exact=["vm"], contains=["vm name", "name"])
    col_cpu = _find_column(headers, exact=["cpus"], contains=["cpu"])
    col_mem = _find_column(headers, exact=["memory"], contains=["memory"])
    col_disk = _find_column(headers, exact=["provisioned mib", "provisioned mb"],
                            contains=["provisioned", "in use"])
    col_power = _find_column(headers, exact=["powerstate"], contains=["power"])
    col_tpl = _find_column(headers, exact=["template"], contains=["template"])

    if col_cpu is None or col_mem is None:
        raise ValueError(
            f"Could not find CPU/Memory columns in sheet '{ws.title}'. "
            f"Headers seen: {headers[:12]}...")

    mem_header = headers[col_mem]
    disk_header = headers[col_disk] if col_disk is not None else ""

    vms: List[Dict[str, Any]] = []
    skipped = {"poweredoff": 0, "templates": 0, "no_spec": 0}

    for row in rows:
        if row is None or all(c is None for c in row):
            continue

        def cell(idx):
            return row[idx] if idx is not None and idx < len(row) else None

        if not include_templates and col_tpl is not None:
            tpl = str(cell(col_tpl) or "").strip().lower()
            if tpl in ("true", "1", "yes"):
                skipped["templates"] += 1
                continue

        if not include_off and col_power is not None:
            power = str(cell(col_power) or "").strip().lower()
            if power and "off" in power:
                skipped["poweredoff"] += 1
                continue

        vcpu = _to_number(cell(col_cpu))
        ram_gib = _unit_to_gib(_to_number(cell(col_mem)), mem_header)
        if not vcpu or not ram_gib:
            skipped["no_spec"] += 1
            continue

        disk_gib = _unit_to_gib(_to_number(cell(col_disk)), disk_header) if col_disk is not None else None

        vms.append({
            "name": str(cell(col_name) or "?").strip() if col_name is not None else "?",
            "vcpu": int(round(vcpu)),
            "ram_gib": round(ram_gib, 1),
            "disk_gib": round(disk_gib, 1) if disk_gib else None,
        })

    return vms, skipped


# ── Pricing & mapping ──────────────────────────────────────────────────────

def map_vms(vms, args):
    """Attach an Azure flavor, disk tier and live prices to each VM. Prices are
    cached per SKU / per disk tier so a 500-VM sheet makes only a handful of
    API calls."""
    catalog = load_catalog()
    tiers = load_disk_tiers()
    vm_price_cache: Dict[str, Dict[str, Optional[float]]] = {}
    disk_price_cache: Dict[str, Optional[float]] = {}

    for vm in vms:
        flavor = pick_flavor(catalog, vm["vcpu"], vm["ram_gib"])
        vm["flavor"] = flavor
        vm["undersized"] = bool(flavor) and (flavor["vcpu"] < vm["vcpu"] or flavor["ram_gib"] < vm["ram_gib"])

        sku = flavor["sku"] if flavor else None
        if sku and sku not in vm_price_cache:
            vm_price_cache[sku] = organize_vm_prices(query_vm_prices(sku, args.region, args.currency))
        vm["prices"] = vm_price_cache.get(sku, {})

        if vm["disk_gib"]:
            info = disk_tier_for_size(vm["disk_gib"], args.disk_type, tiers)
            vm["disk_tier"] = info
            key = info["sku_name"]
            if key not in disk_price_cache:
                disk_price_cache[key] = query_disk_price(
                    info["sku_name"], info["product_name"], args.region, args.currency)
            vm["disk_monthly"] = disk_price_cache[key]
        else:
            vm["disk_tier"] = None
            vm["disk_monthly"] = None
    return vms


# ── Rendering ──────────────────────────────────────────────────────────────

def money(v, sym, dp=2):
    return "N/A" if v is None else f"{sym}{v:,.{dp}f}"


def render(vms, skipped, args, sym):
    print(f"# Azure sizing from RVTools — {os.path.basename(args.file)}\n")
    print(f"Target region **{args.region}** · {args.currency} · OS **{args.os}** · "
          f"strategy **lift-and-shift** (meet-or-exceed source spec)\n")

    note_bits = [f"{len(vms)} VMs sized"]
    if skipped["poweredoff"]:
        note_bits.append(f"{skipped['poweredoff']} powered-off skipped")
    if skipped["templates"]:
        note_bits.append(f"{skipped['templates']} templates skipped")
    if skipped["no_spec"]:
        note_bits.append(f"{skipped['no_spec']} without specs skipped")
    print("> " + " · ".join(note_bits) + "\n")

    if not vms:
        print("No VMs to size. Try `--include-poweredoff` / `--include-templates`, "
              "or check the sheet has CPU/Memory columns.")
        return False

    os_key = args.os

    def vm_monthly(vm, hourly_key):
        h = vm["prices"].get(hourly_key)
        return None if h is None else h * HOURS_PER_MONTH

    # --- Rollup totals ---
    payg_vm = sum((vm_monthly(vm, os_key) or 0) for vm in vms)
    res_vm = sum((vm_monthly(vm, "reserved_1yr") or 0) for vm in vms)
    disk_total = sum((vm["disk_monthly"] or 0) for vm in vms)
    payg_total = payg_vm + disk_total
    res_total = res_vm + disk_total
    total_vcpu = sum(vm["vcpu"] for vm in vms)
    total_ram = sum(vm["ram_gib"] for vm in vms)

    print("## Summary\n")
    print("| Metric | Value |")
    print("|--------|-------|")
    print(f"| VMs | {len(vms)} |")
    print(f"| Source vCPU / RAM (allocated) | {total_vcpu} vCPU / {total_ram:,.0f} GiB |")
    print(f"| Compute PAYG | {money(payg_vm, sym)} /mo · {money(payg_vm * 12, sym)} /yr |")
    print(f"| Compute 1yr Reserved | {money(res_vm, sym)} /mo · {money(res_vm * 12, sym)} /yr |")
    print(f"| Managed disk (no reservation discount) | {money(disk_total, sym)} /mo |")
    print(f"| **Total PAYG** | **{money(payg_total, sym)} /mo · {money(payg_total * 12, sym)} /yr** |")
    print(f"| **Total 1yr Reserved** | **{money(res_total, sym)} /mo · {money(res_total * 12, sym)} /yr** |")
    if payg_total:
        save = (payg_total - res_total) / payg_total * 100
        print(f"| Reserved saving | -{save:.0f}% vs PAYG |")

    # --- By-flavor rollup ---
    by_flavor: Dict[str, Dict[str, Any]] = {}
    for vm in vms:
        sku = vm["flavor"]["sku"] if vm["flavor"] else "—"
        agg = by_flavor.setdefault(sku, {"count": 0, "flavor": vm["flavor"],
                                         "unit": vm_monthly(vm, os_key)})
        agg["count"] += 1
    print("\n## By Azure flavor\n")
    print("| Azure SKU | Type | vCPU | RAM | Count | Unit PAYG /mo | Subtotal /mo |")
    print("|-----------|------|------|-----|-------|---------------|--------------|")
    for sku, agg in sorted(by_flavor.items(), key=lambda kv: -kv[1]["count"]):
        f = agg["flavor"]
        unit = agg["unit"]
        sub = None if unit is None else unit * agg["count"]
        fam = FAMILY_LABEL.get(f["family"], f["family"]) if f else "?"
        vcpu = f["vcpu"] if f else "?"
        ram = f"{f['ram_gib']} GiB" if f else "?"
        print(f"| {sku} | {fam} | {vcpu} | {ram} | {agg['count']} "
              f"| {money(unit, sym)} | {money(sub, sym)} |")

    # --- Per-VM detail ---
    print("\n## Per-VM mapping\n")
    print("| VM | Src vCPU/RAM | Src disk | → Azure SKU | vCPU/RAM | Disk tier | PAYG /mo | 1yr /mo |")
    print("|----|--------------|----------|-------------|----------|-----------|----------|---------|")
    for vm in vms:
        f = vm["flavor"]
        sku = f["sku"] if f else "—"
        az_spec = f"{f['vcpu']} / {f['ram_gib']}G" if f else "?"
        flag = " ⚠️" if vm["undersized"] else ""
        src_disk = f"{vm['disk_gib']:g}G" if vm["disk_gib"] else "—"
        tier = vm["disk_tier"]["sku_name"] if vm["disk_tier"] else "—"
        vm_m = vm_monthly(vm, os_key)
        res_m = vm_monthly(vm, "reserved_1yr")
        disk_m = vm["disk_monthly"] or 0
        payg = None if vm_m is None else vm_m + disk_m
        resv = None if res_m is None else res_m + disk_m
        print(f"| {vm['name']} | {vm['vcpu']} / {vm['ram_gib']:g}G | {src_disk} "
              f"| {sku}{flag} | {az_spec} | {tier} | {money(payg, sym)} | {money(resv, sym)} |")

    if any(vm["undersized"] for vm in vms):
        print("\n> ⚠️ = no catalog flavor large enough; largest available was used. "
              "Extend `references/vm-catalog.json` (e.g. GPU/large sizes).")

    print("\n> Lift-and-shift: each VM mapped to the cheapest flavor meeting-or-exceeding "
          "its **allocated** vCPU/RAM. RVTools is a point-in-time allocation snapshot "
          "(no performance data, no Azure region) — for right-sizing use Azure Migrate. "
          "Provisioned storage is modelled as one managed disk billed at the next tier up.")
    print(f"\n*Generated {date.today().isoformat()} from Azure Retail Prices API. "
          f"Prices are indicative and subject to change.*")
    return True


def auto_filename(args):
    base = re.sub(r"[^A-Za-z0-9_-]", "", os.path.splitext(os.path.basename(args.file))[0])
    return os.path.join("quotes", f"rvtools-{base}-{args.region}-{date.today().isoformat()}.md")


def main():
    p = argparse.ArgumentParser(description="Batch Azure sizing from an RVTools xlsx export")
    p.add_argument("file", help="Path to the RVTools .xlsx export")
    p.add_argument("--region", "-r", default="francecentral", help="Target Azure region (default: francecentral)")
    p.add_argument("--currency", "-c", default="EUR", help="Currency code (default: EUR)")
    p.add_argument("--os", default="linux", choices=["linux", "windows"],
                   help="OS for the headline totals (default: linux)")
    p.add_argument("--disk-type", default="premium-ssd",
                   choices=["premium-ssd", "standard-ssd", "standard-hdd"],
                   help="Managed disk type for all VMs (default: premium-ssd)")
    p.add_argument("--sheet", default=None, help="Worksheet name (default: auto-detect vInfo)")
    p.add_argument("--include-poweredoff", action="store_true", help="Include powered-off VMs")
    p.add_argument("--include-templates", action="store_true", help="Include VM templates")
    p.add_argument("--output", "-o", nargs="?", const="auto", default=None, metavar="PATH",
                   help="Write the report to Markdown. Bare flag auto-names it under quotes/; "
                        "pass a PATH to choose the location.")
    args = p.parse_args()

    if not os.path.isfile(args.file):
        print(f"File not found: {args.file}", file=sys.stderr)
        sys.exit(1)

    sym = currency_symbol(args.currency)
    try:
        vms, skipped = parse_rvtools(args.file, args.sheet,
                                     args.include_poweredoff, args.include_templates)
    except Exception as e:
        print(f"Failed to parse RVTools file: {e}", file=sys.stderr)
        sys.exit(1)

    vms = map_vms(vms, args)

    buf = io.StringIO()
    with redirect_stdout(buf):
        ok = render(vms, skipped, args, sym)
    text = buf.getvalue()
    print(text, end="" if text.endswith("\n") else "\n")

    if args.output and ok:
        path = auto_filename(args) if args.output == "auto" else args.output
        parent = os.path.dirname(path)
        if parent:
            os.makedirs(parent, exist_ok=True)
        with open(path, "w", encoding="utf-8") as f:
            f.write(text)
        print(f"\n✅ Written to {path}", file=sys.stderr)


if __name__ == "__main__":
    main()
