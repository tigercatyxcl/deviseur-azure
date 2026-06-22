#!/usr/bin/env python3
"""
Export a multi-VM (fleet) Azure quote to a multi-sheet Excel workbook.

Builds, from live Azure Retail Prices, an .xlsx with:
  * Selection   — the per-group source spec → Azure SKU mapping
  * Per-Group   — monthly cost per group across every commitment option
  * Fleet Total — the whole-fleet monthly + annual rollup
  * TCO         — cumulative cost over 1 / 2 / 3 year ownership horizons

Windows groups carry two reserved figures: **no AHB** (the Windows Server
licence stays at PAYG on top of the reserved compute) and **with AHB** (Azure
Hybrid Benefit — only the reserved compute is billed). Azure VM reservations
exist only in 1yr and 3yr terms; the 2-year TCO column is a *time horizon*
(a 1yr reservation renewed), not a 2-year reservation product.

The fleet is defined on the command line with one repeatable --group per
distinct spec, each "SPEC,OS,QTY[,SKU]". The SKU is auto-picked via the sizing
rule (RAM meet-or-exceed, vCPU floor, D-family preferred) unless pinned.

Usage:
    python3 scripts/export_fleet_xlsx.py \
        --group 4U8G,windows,10 --group 3U6G,linux,10 --group 5U11G,linux,10
    # pin a SKU and choose region / output path:
    python3 scripts/export_fleet_xlsx.py --group 8U64G,linux,5,E8s_v5 \
        --region westeurope --output /tmp/quote.xlsx
"""

import argparse
import os
import re
import string
import sys
from datetime import date

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from azure_lib import (  # noqa: E402
    query_vm_prices, organize_vm_prices, currency_symbol, HOURS_PER_MONTH,
    load_catalog, pick_flavor, normalize_sku,
)

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── Fleet definition ────────────────────────────────────────────────────────
# Groups are supplied on the command line via repeatable --group arguments:
#   --group "4U8G,windows,10"        (source spec, OS, qty; SKU auto-picked)
#   --group "5U11G,linux,10,D8s_v5"  (4th field pins an explicit SKU)
# See parse_group() / build_groups() below.

# Commitment keys carried through every table (None = not available).
COMMITMENTS = [
    "payg", "spot", "ri1y_noahb", "ri1y_ahb", "ri3y_noahb", "ri3y_ahb",
]
COMMIT_LABEL = {
    "payg": "PAYG (on-demand)",
    "spot": "Spot",
    "ri1y_noahb": "1yr Reserved (Win no AHB)",
    "ri1y_ahb": "1yr Reserved (Win AHB)",
    "ri3y_noahb": "3yr Reserved (Win no AHB)",
    "ri3y_ahb": "3yr Reserved (Win AHB)",
}


SPEC_RE = re.compile(r"^\s*(\d+)\s*[uUcC]\s*(\d+(?:\.\d+)?)\s*[gG]?\s*$")


def parse_spec(spec):
    """Parse a shorthand spec like '4U8G' / '3c6g' / '5U11' into (vcpu, ram_gib)."""
    m = SPEC_RE.match(spec)
    if not m:
        raise ValueError(
            f"Bad spec '{spec}'. Use <vcpu>U<ram>G, e.g. 4U8G or 5U11G.")
    return int(m.group(1)), float(m.group(2))


def parse_group(raw):
    """Parse one --group value 'SPEC,OS,QTY[,SKU]' into a partial group dict."""
    parts = [p.strip() for p in raw.split(",")]
    if len(parts) < 3:
        raise ValueError(
            f"Bad --group '{raw}'. Expected 'SPEC,OS,QTY[,SKU]', e.g. '4U8G,windows,10'.")
    spec, os_raw, qty_raw = parts[0], parts[1].lower(), parts[2]
    if os_raw not in ("linux", "windows"):
        raise ValueError(f"OS must be 'linux' or 'windows', got '{parts[1]}' in '{raw}'.")
    try:
        qty = int(qty_raw)
    except ValueError:
        raise ValueError(f"QTY must be an integer, got '{qty_raw}' in '{raw}'.")
    if qty < 1:
        raise ValueError(f"QTY must be >= 1 in '{raw}'.")
    vcpu, ram = parse_spec(spec)
    sku = parts[3] if len(parts) >= 4 and parts[3] else None
    return {"src": spec.upper(), "os": os_raw, "qty": qty,
            "vcpu_src": vcpu, "ram_src": ram, "sku_override": sku}


def build_groups(raw_groups):
    """Turn raw --group strings into full group dicts, applying the sizing rule
    (RAM meet-or-exceed, vCPU floor, D-family preferred) to pick a SKU unless one
    is given explicitly. Labels are auto-assigned A, B, C ..."""
    catalog = load_catalog()
    by_sku = {s["sku"]: s for s in catalog}
    labels = list(string.ascii_uppercase)
    groups = []
    for i, raw in enumerate(raw_groups):
        g = parse_group(raw)
        if g["sku_override"]:
            sku = normalize_sku(g["sku_override"])
            cat = by_sku.get(sku)
            vcpu = cat["vcpu"] if cat else g["vcpu_src"]
            ram = cat["ram_gib"] if cat else g["ram_src"]
        else:
            flavor = pick_flavor(catalog, g["vcpu_src"], g["ram_src"])
            if not flavor:
                raise ValueError(f"No catalog flavor for spec '{g['src']}'.")
            sku, vcpu, ram = flavor["sku"], flavor["vcpu"], flavor["ram_gib"]
        groups.append({
            "label": labels[i] if i < len(labels) else f"G{i + 1}",
            "src": g["src"], "os": g["os"], "qty": g["qty"],
            "sku": sku, "vcpu": vcpu, "ram": ram,
        })
    return groups


def group_monthly(g, region, currency):
    """Return {commitment: monthly_total_for_group} (None where unavailable)."""
    pr = organize_vm_prices(query_vm_prices(g["sku"], region, currency))
    qty, mo = g["qty"], HOURS_PER_MONTH

    def m(hourly):
        return None if hourly is None else hourly * mo * qty

    if g["os"] == "windows":
        lic = ((pr["windows"] - pr["linux"])
               if pr["windows"] is not None and pr["linux"] is not None else 0.0)

        def wl(hourly):  # reserved compute + licence at PAYG (no AHB)
            return None if hourly is None else hourly + lic

        return {
            "payg": m(pr["windows"]),
            "spot": None,  # Spot is Linux-only
            "ri1y_noahb": m(wl(pr["reserved_1yr"])),
            "ri1y_ahb": m(pr["reserved_1yr"]),
            "ri3y_noahb": m(wl(pr["reserved_3yr"])),
            "ri3y_ahb": m(pr["reserved_3yr"]),
        }
    # Linux: AHB columns mirror the base reserved rate (no licence involved).
    return {
        "payg": m(pr["linux"]),
        "spot": m(pr["spot"]),
        "ri1y_noahb": m(pr["reserved_1yr"]),
        "ri1y_ahb": m(pr["reserved_1yr"]),
        "ri3y_noahb": m(pr["reserved_3yr"]),
        "ri3y_ahb": m(pr["reserved_3yr"]),
    }


def add(a, b):
    """Sum two optional numbers; None only if BOTH are None."""
    if a is None and b is None:
        return None
    return (a or 0) + (b or 0)


# ── Styling helpers ─────────────────────────────────────────────────────────
HDR_FILL = PatternFill("solid", fgColor="1F4E78")
HDR_FONT = Font(bold=True, color="FFFFFF")
TITLE_FONT = Font(bold=True, size=13, color="1F4E78")
BOLD = Font(bold=True)
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
MONEY_FMT = '#,##0.00 "€"'


def style_header(ws, row, ncols):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HDR_FILL
        cell.font = HDR_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER


def autofit(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w


def write_row(ws, row, values, money_cols=(), bold=False):
    for i, v in enumerate(values, 1):
        cell = ws.cell(row=row, column=i, value=v)
        cell.border = BORDER
        if bold:
            cell.font = BOLD
        if i in money_cols and isinstance(v, (int, float)):
            cell.number_format = MONEY_FMT
            cell.alignment = Alignment(horizontal="right")


def build(path, groups, region, currency):
    sym = currency_symbol(currency)  # noqa: F841 (kept for future symbol use)
    for g in groups:
        g["monthly"] = group_monthly(g, region, currency)

    # Fleet monthly per commitment
    fleet = {k: None for k in COMMITMENTS}
    for k in COMMITMENTS:
        for g in groups:
            fleet[k] = add(fleet[k], g["monthly"][k])

    wb = openpyxl.Workbook()

    # ---- Sheet 1: Selection ----
    ws = wb.active
    ws.title = "Selection"
    ws["A1"] = f"Azure Quote — {len(groups)} groups / {sum(g['qty'] for g in groups)} VMs @ {region} ({currency})"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = (f"Generated {date.today().isoformat()} · live Azure Retail Prices · "
                f"sizing rule: RAM meets-or-exceeds, vCPU floors to nearest size, D-family preferred")
    ws["A2"].font = Font(italic=True, color="808080")
    hdr = ["Group", "Source spec", "OS", "Qty", "Azure SKU", "vCPU", "RAM (GiB)"]
    ws.append([])
    ws.append(hdr)
    style_header(ws, ws.max_row, len(hdr))
    for g in groups:
        write_row(ws, ws.max_row + 1,
                  [g["label"], g["src"], g["os"].capitalize(), g["qty"],
                   g["sku"], g["vcpu"], g["ram"]])
    autofit(ws, [8, 14, 10, 6, 20, 7, 11])

    # ---- Sheet 2: Per-Group monthly ----
    ws = wb.create_sheet("Per-Group")
    ws["A1"] = "Monthly cost per group — all commitment options (€/month)"
    ws["A1"].font = TITLE_FONT
    hdr = ["Group", "OS", "Qty", "SKU"] + [COMMIT_LABEL[k] for k in COMMITMENTS]
    ws.append([])
    ws.append(hdr)
    style_header(ws, ws.max_row, len(hdr))
    money_cols = tuple(range(5, 5 + len(COMMITMENTS)))
    for g in groups:
        row = [g["label"], g["os"].capitalize(), g["qty"], g["sku"]]
        row += [g["monthly"][k] for k in COMMITMENTS]
        write_row(ws, ws.max_row + 1, row, money_cols=money_cols)
    # fleet subtotal
    total_row = ["TOTAL", "", sum(g["qty"] for g in groups), ""] + [fleet[k] for k in COMMITMENTS]
    write_row(ws, ws.max_row + 1, total_row, money_cols=money_cols, bold=True)
    autofit(ws, [8, 10, 6, 20] + [24] * len(COMMITMENTS))

    # ---- Sheet 3: Fleet Total (monthly + annual) ----
    ws = wb.create_sheet("Fleet Total")
    ws["A1"] = f"Whole-fleet rollup — {sum(g['qty'] for g in groups)} VMs (€)"
    ws["A1"].font = TITLE_FONT
    hdr = ["Commitment", "Total / month", "Total / year", "vs PAYG"]
    ws.append([])
    ws.append(hdr)
    style_header(ws, ws.max_row, len(hdr))
    payg_y = (fleet["payg"] or 0) * 12
    for k in COMMITMENTS:
        mo = fleet[k]
        if mo is None:
            write_row(ws, ws.max_row + 1, [COMMIT_LABEL[k], "N/A", "N/A", "N/A"])
            continue
        yr = mo * 12
        vs = "—" if k == "payg" else f"-{(payg_y - yr) / payg_y * 100:.0f}%"
        write_row(ws, ws.max_row + 1, [COMMIT_LABEL[k], mo, yr, vs],
                  money_cols=(2, 3), bold=(k == "payg"))
    ws.append([])
    note = ws.cell(row=ws.max_row + 1, column=1,
                   value="Spot shown for Linux groups only (B+C); Windows has no Spot rate. "
                         "Windows reserved given both with/without Azure Hybrid Benefit.")
    note.font = Font(italic=True, color="808080")
    autofit(ws, [30, 16, 16, 10])

    # ---- Sheet 4: TCO over 1/2/3 year horizons ----
    ws = wb.create_sheet("TCO")
    ws["A1"] = "Total Cost of Ownership — cumulative spend by horizon (€)"
    ws["A1"].font = TITLE_FONT
    ws["A2"] = ("Azure VM reservations exist only in 1yr & 3yr terms. The 2-year column is a "
                "time horizon (1yr reservation renewed), not a 2-year reservation product. "
                "A 3yr reservation requires a full 3-year commitment.")
    ws["A2"].font = Font(italic=True, color="808080")
    hdr = ["Strategy", "€/month", "1-year TCO", "2-year TCO", "3-year TCO"]
    ws.append([])
    ws.append(hdr)
    style_header(ws, ws.max_row, len(hdr))
    for k in COMMITMENTS:
        if k == "spot":
            continue  # spot is interruptible, not a TCO commitment baseline
        mo = fleet[k]
        if mo is None:
            continue
        write_row(ws, ws.max_row + 1,
                  [COMMIT_LABEL[k], mo, mo * 12, mo * 24, mo * 36],
                  money_cols=(2, 3, 4, 5), bold=(k == "payg"))
    autofit(ws, [30, 14, 16, 16, 16])

    parent = os.path.dirname(path)
    if parent:
        os.makedirs(parent, exist_ok=True)
    wb.save(path)
    return fleet


def main():
    p = argparse.ArgumentParser(
        description="Export a multi-group fleet Azure quote to Excel",
        epilog="Example: export_fleet_xlsx.py "
               "--group 4U8G,windows,10 --group 3U6G,linux,10 --group 5U11G,linux,10",
        formatter_class=argparse.RawDescriptionHelpFormatter)
    p.add_argument("--group", "-g", action="append", dest="groups", metavar="SPEC,OS,QTY[,SKU]",
                   help="A VM group: source spec, OS (linux/windows), quantity, and an "
                        "optional explicit SKU. Repeat --group for each distinct group. "
                        "SKU is auto-picked from the sizing rule when omitted.")
    p.add_argument("--region", "-r", default="francecentral", help="Azure region (default: francecentral)")
    p.add_argument("--currency", "-c", default="EUR", help="Currency code (default: EUR)")
    p.add_argument("--output", "-o", default=None, help="Output .xlsx path (default: auto under quotes/)")
    args = p.parse_args()

    if not args.groups:
        p.error("at least one --group is required, e.g. --group 4U8G,windows,10")

    try:
        groups = build_groups(args.groups)
    except ValueError as e:
        p.error(str(e))

    output = args.output or os.path.join(
        "quotes", f"azure-fleet-quote-{args.region}-{date.today().isoformat()}.xlsx")
    fleet = build(output, groups, args.region, args.currency)

    print(f"✅ Wrote {output}")
    print(f"Fleet: {len(groups)} groups / {sum(g['qty'] for g in groups)} VMs @ "
          f"{args.region} ({args.currency})")
    for g in groups:
        print(f"  {g['label']}: {g['src']:>7} {g['os']:<7} ×{g['qty']:<3} → "
              f"{g['sku']} ({g['vcpu']}vCPU/{g['ram']}GiB)")
    print("Monthly totals:")
    for k in COMMITMENTS:
        v = fleet[k]
        print(f"  {COMMIT_LABEL[k]:32s} {'N/A' if v is None else f'€{v:,.2f}'}")


if __name__ == "__main__":
    main()
